#!/usr/bin/env python3
"""Ingest NHSE 'Ambulance Handover Times by Acute Trust' (DAC handovers) files
and fold handover data into data/uec.json (run ingest_ae.py first).

Usage:
    python3 ingest/ingest_handover.py --timeseries <ts.xlsx> <apr26.xlsx> [<apr25.xlsx>...] [--provider RX1]

Monthly snapshot workbooks ('All handovers' sheet): counts by threshold
(>15/>30/>45/>60 min), mean handover time, and crew hours lost beyond 30
minutes. Timeseries workbook: monthly mean handover time and handover counts
per trust from Oct-2023.
"""
import datetime
import json
import os
import sys
import openpyxl


def to_minutes(v):
    if v in (None, "", "-"):
        return None
    if isinstance(v, datetime.time):
        return round(v.hour * 60 + v.minute + v.second / 60, 1)
    if isinstance(v, datetime.timedelta):
        return round(v.total_seconds() / 60, 1)
    # unformatted Excel time serial (fraction of a day)
    if isinstance(v, (int, float)) and 0 < v < 1:
        return round(v * 1440, 1)
    s = str(v).strip()
    parts = s.split(":")
    if len(parts) == 3:
        return round(int(parts[0]) * 60 + int(parts[1]) + float(parts[2]) / 60, 1)
    return None


def read_snapshot(path, provider):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["All handovers"]
    ws.reset_dimensions()
    period, has45 = None, None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        c = list(row) + [None] * 24
        if i == 1 and isinstance(c[0], datetime.datetime):
            period = c[0].strftime("%b-%y")
        # the 2024 edition has no 'Over 45 minutes' column — detect from the header
        if has45 is None and str(c[0] or "") == "Region":
            has45 = "45" in str(c[8] or "")
        if str(c[1] or "").strip() == provider:
            wb.close()
            # column layout after the >30 column shifts by one when >45 is present
            o = 1 if has45 else 0
            known = float(c[5] or 0)
            rec = {
                "handovers": int(c[10 + o] or 0), "known": int(known),
                "over15": int(c[6] or 0), "over30": int(c[7] or 0),
                "over45": int(c[8] or 0) if has45 else None, "over60": int(c[8 + o] or 0),
                "meanMin": to_minutes(c[13 + o]),
                "hoursLostOver30": round(float(c[14 + o] or 0)),
            }
            for k in ("over15", "over30", "over60"):
                rec[f"{k}Pct"] = round(100 * rec[k] / known, 1) if known else None
            return period or os.path.basename(path), rec
    raise SystemExit(f"{provider} not found in {path}")


def read_timeseries(path, provider):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet, key in [("Average handover time", "meanMin"), ("Counts", "handovers")]:
        ws = wb[sheet]
        ws.reset_dimensions()
        months, rx = [], None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            c = list(row)
            if i == 3:
                # the final columns may be plain month-name strings ('Apr', 'May')
                # continuing on from the last dated column
                months, last = [], None
                ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                for j, v in enumerate(c):
                    if isinstance(v, datetime.datetime):
                        last = (v.year, v.month)
                        months.append((j, f"{v.year}-{v.month:02d}"))
                    elif last and str(v).strip()[:3] in ABBR:
                        y, m = last
                        m += 1
                        if m > 12: y, m = y + 1, 1
                        if ABBR[m - 1] == str(v).strip()[:3]:
                            last = (y, m)
                            months.append((j, f"{y}-{m:02d}"))
            if str(c[2] if len(c) > 2 else "").strip() == provider:
                rx = c
                break
        vals = {ym: rx[j] for j, ym in months}
        if not out:
            out = [{"ym": ym} for _, ym in months]
        for rec in out:
            v = vals.get(rec["ym"])
            rec[key] = to_minutes(v) if key == "meanMin" else (int(v) if isinstance(v, (int, float)) else None)
    wb.close()
    return [r for r in out if r.get("meanMin") is not None]


def main():
    provider = sys.argv[sys.argv.index("--provider") + 1] if "--provider" in sys.argv else "RX1"
    ts_path = sys.argv[sys.argv.index("--timeseries") + 1]
    skip = {provider, ts_path}
    snaps = [a for a in sys.argv[1:] if not a.startswith("--") and a not in skip]

    months = dict(sorted((read_snapshot(p, provider) for p in snaps),
                         key=lambda t: (t[0][-2:], "JanFebMarAprMayJunJulAugSepOctNovDec".find(t[0][:3]))))
    series = read_timeseries(ts_path, provider)

    for label, m in months.items():
        print(f"  {label}: {m['handovers']:,} handovers | mean {m['meanMin']}min | "
              f">30min {m['over30Pct']}% | >60min {m['over60Pct']}% | {m['hoursLostOver30']:,}h lost")
    peak = max(series, key=lambda r: r["meanMin"])
    print(f"  timeseries: {len(series)} months {series[0]['ym']} → {series[-1]['ym']} | "
          f"peak mean {peak['meanMin']}min ({peak['ym']}) | latest {series[-1]['meanMin']}min")

    out_path = os.path.join(os.path.dirname(__file__), "..", "data", "uec.json")
    uec = json.load(open(out_path))
    uec["handover"] = {
        "note": ("Ambulance Handover Times by Acute Trust (NHSE management information, "
                 "'All handovers'). National standard: handover within 15 minutes; hours "
                 "lost counted beyond 30 minutes. Mean is over handovers with known time."),
        "months": months,
        "timeseries": series,
    }
    json.dump(uec, open(out_path, "w"), indent=2)
    print(f"updated {os.path.normpath(out_path)}")


if __name__ == "__main__":
    main()
