#!/usr/bin/env python3
"""Ingest NHSE Cancer Waiting Times provider data and build data/cancer.json.

Handles both published formats ("because NHS"):
  - Monthly Combined CSV (2026+): long format, Basis/Org_Code/Standard_or_Item/
    Cancer_Type/Referral_Route_or_Stage/Treatment_Modality/Total/Within/After.
  - CWT CRS Provider Workbook (.xlsx, 2024/25): one sheet per standard
    ('28-DAY FDS (ALL ROUTES)', '31-DAY (ALL CANCER)', '62-DAY (ALL CANCER)',
    plus BY CANCER sheets), filtered to the ALL ROUTES part.

Usage:
    python3 ingest/ingest_cwt.py <latest.(csv|xlsx)> [--prior <older.(csv|xlsx)>]... [--provider RX1]

Outputs per snapshot the three standards (FDS 28-day, 31-day, 62-day) as
{total, within, pct}, tumour-site detail from the latest snapshot, and demand
growth calibrated from the EARLIEST adjacent pair (pre-EPR, as per RTT/DM01).
"""
import csv
import json
import os
import re
import sys


def read_combined_csv(path, provider):
    out = {"fds": None, "d31": None, "d62": None, "sites": {}}
    period = None
    with open(path, encoding="utf-8-sig") as f:
        for d in csv.DictReader(f):
            if d["Basis"] != "Provider" or d["Org_Code"].strip() != provider:
                continue
            std = d["Standard_or_Item"]
            ct = d["Cancer_Type"].strip()
            route = d["Referral_Route_or_Stage"].strip().upper()
            mod = d["Treatment_Modality"].strip().upper()
            tot = float(d["Total"] or 0); win = float(d["Within"] or 0)
            rec = {"total": tot, "within": win, "pct": round(100 * win / tot, 1) if tot else None}
            allroute = route in ("ALL ROUTES", "ALL STAGES") and mod in ("", "ALL MODALITIES")
            if not allroute:
                continue
            if ct == "ALL CANCERS":
                if std == "FDS": out["fds"] = rec
                elif std == "31D": out["d31"] = rec
                elif std == "62D": out["d62"] = rec
            elif std in ("FDS", "31D", "62D"):
                key = {"FDS": "fds", "31D": "d31", "62D": "d62"}[std]
                out["sites"].setdefault(ct, {})[key] = rec
    # the combined CSV carries no period column — derive 'Apr-26' from the filename
    m = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-?(\d{2})", os.path.basename(path))
    return out, period or (f"{m.group(1)}-{m.group(2)}" if m else os.path.basename(path))


def read_crs_workbook(path, provider):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"fds": None, "d31": None, "d62": None, "sites": {}}
    period = None

    def rows_of(sheet):
        ws = wb[sheet]
        # some CWT workbooks carry a broken dimension record (A1:A1), which makes
        # read_only iteration yield a single empty row — recompute from the data
        ws.reset_dimensions()
        for row in ws.iter_rows(values_only=True):
            yield [v if v is not None else "" for v in row]

    # sheet -> (standard key, has cancer-type column)
    plans = [("28-DAY FDS (ALL ROUTES)", "fds", False), ("31-DAY (ALL CANCER)", "d31", False),
             ("62-DAY (ALL CANCER)", "d62", False), ("31-DAY (BY CANCER)", "d31", True),
             ("62-DAY (BY CANCER)", "d62", True)]
    for sheet, key, by_cancer in plans:
        if sheet not in wb.sheetnames:
            continue
        for row in rows_of(sheet):
            cells = [str(c).strip() for c in row]
            if period is None and len(cells) > 0 and cells[0][:3] in ("Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec") and "-" in cells[0]:
                period = cells[0]
            if len(cells) < 8 or cells[1] != provider:
                continue
            route = cells[3].upper()
            if route not in ("ALL ROUTES", "ALL STAGES"):
                continue
            # FDS ALL-ROUTES: totals at cols 4/5. The 31/62-day sheets (both ALL
            # CANCER and BY CANCER) carry a CANCER TYPE column at 4, totals at 5/6.
            has_type_col = key != "fds"
            ct = cells[4] if has_type_col else "ALL CANCERS"
            ti = 5 if has_type_col else 4
            tot, win = float(row[ti] or 0), float(row[ti + 1] or 0)
            rec = {"total": tot, "within": win, "pct": round(100 * win / tot, 1) if tot else None}
            if by_cancer:
                out["sites"].setdefault(ct, {})[key] = rec
            elif ct.upper() == "ALL CANCERS":
                out[key] = rec
    return out, period or os.path.basename(path)


def read_any(path, provider):
    if path.lower().endswith(".csv"):
        return read_combined_csv(path, provider)
    return read_crs_workbook(path, provider)


def main():
    src = sys.argv[1]
    provider = sys.argv[sys.argv.index("--provider") + 1] if "--provider" in sys.argv else "RX1"
    prior_srcs = [sys.argv[i + 1] for i, a in enumerate(sys.argv) if a == "--prior"]

    cur, period = read_any(src, provider)
    priors = [read_any(p, provider) for p in prior_srcs]
    priors.sort(key=lambda t: t[1])

    chain = [p[0] for p in priors] + [cur]
    labels = [p[1] for p in priors] + [period]
    for lab, snap in zip(labels, chain):
        f, t1, t2 = snap["fds"], snap["d31"], snap["d62"]
        print(f"  {lab}: FDS {f['pct']}% ({f['total']:.0f}) | 31D {t1['pct']}% ({t1['total']:.0f}) | 62D {t2['pct']}% ({t2['total']:.0f})")

    growth, calNote = 0, "no prior data"
    if len(chain) >= 2:
        g = []
        for i in range(len(chain) - 1):
            g.append(round(100 * (chain[i + 1]["fds"]["total"] / chain[i]["fds"]["total"] - 1), 1))
            print(f"  pair {labels[i]} -> {labels[i+1]}: FDS volume growth {g[-1]:+.1f}%/yr")
        growth = g[0]
        calNote = f"FDS-volume growth from earliest pair ({labels[0]} vs {labels[1]})"

    out = {
        "_provenance": {
            "provider": provider, "period": period, "sources": labels,
            "dataQuality": ("Seeded from published CWT provider data (combined CSV 2026; CRS "
                            "workbooks 2024/25 — two formats, one statistic). April volumes are "
                            f"bank-holiday depressed. {calNote}; pre-EPR pair preferred as per "
                            "RTT/DM01. Published CWT has no PTL backlog census — the model works "
                            "on treated/seen volumes and timeliness shares; trust PTL data would "
                            "deepen it."),
        },
        "standards": {
            "fds": {"name": "Faster Diagnosis (28-day)", "targetPct": 80,
                    "milestones": [{"ym": "2027-04", "pct": 80}],
                    "current": cur["fds"], "history": {l: c["fds"] for l, c in zip(labels, chain)}},
            "d31": {"name": "31-day decision-to-treat", "targetPct": 96,
                    "milestones": [{"ym": "2027-04", "pct": 96}],
                    "current": cur["d31"], "history": {l: c["d31"] for l, c in zip(labels, chain)}},
            "d62": {"name": "62-day referral-to-treatment", "targetPct": 85, "interimPct": 70,
                    "milestones": [{"ym": "2027-04", "pct": 70}, {"ym": "2029-04", "pct": 85}],
                    "current": cur["d62"], "history": {l: c["d62"] for l, c in zip(labels, chain)}},
        },
        "milestoneNote": "milestones are modelling assumptions: FDS 80% and 31D 96% by Apr-27; 62D interim 70% by Apr-27, constitutional 85% by Apr-29 — aligned to the RTT trajectory shape",
        "levers": {"demandGrowthPctYr": growth},
        "sites": cur["sites"],
    }
    path = os.path.join(os.path.dirname(__file__), "..", "data", "cancer.json")
    json.dump(out, open(path, "w"), indent=2)
    print(f"wrote {os.path.normpath(path)} ({len(cur['sites'])} tumour sites)")


if __name__ == "__main__":
    main()
